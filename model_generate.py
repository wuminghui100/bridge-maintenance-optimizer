import numpy as np
import os, math, random
from openpyxl import load_workbook
from mpl_toolkits.mplot3d import Axes3D
import tensorflow as tf
import matplotlib.pyplot as plt
import copy

def not_number(s):
    try:
        float(s)
        return False
    except ValueError:
        pass
 
    try:
        import unicodedata
        unicodedata.numeric(s)
        return False
    except (TypeError, ValueError):
        pass
 
    return True

def max_row(sheet):
    for i in range(1, sheet.max_row+2):
        if sheet.cell(i, 1).value == None:
            return i-1

def processa(num, size):
    # change number num into vector size*1
    vec = np.zeros(shape=[1, size])
    vec[0,num] = 1
    return vec

def deprocessa(vec, size):
    # change vector into int number
    for i in range(size):
        if vec[i]==1:
            return i
    return None

def addData(dataIn, labelIn, num):
    #add several types of data according to domain knowledge; each type has num data
    #1: state=0+action=0 -> state'=0
    #2~4: state=4+action=[1,2,3] -> state'=4
    #5: state=3+action=[3] -> state'=4
    datasize, _ = np.shape(dataIn)
    dataMat = np.zeros(shape=[datasize+5*num, 19], dtype=np.int32)
    labelMat = np.zeros(shape=[datasize+5*num, 5], dtype=np.int32)
    dataMat[0:datasize, :] = dataIn
    labelMat[0:datasize, :] = labelIn
    dataAdd = np.zeros(shape=[1, 9], dtype=np.int32)
    labelAdd = np.zeros(shape=[1, 5], dtype=np.int32)
    ageAdd = np.zeros(shape=[num, 10], dtype=np.int32)
    for i in range(num):
        ageAdd[i,:] = processa(i%10,10)
    #type 1
    dataAdd[0,0:5] = processa(0, 5)
    dataAdd[0,5:9] = processa(0, 4)
    labelAdd = processa(0, 5)
    dataMat[datasize:datasize+num,0:9] = dataAdd
    dataMat[datasize:datasize+num,9:19] = ageAdd
    labelMat[datasize:datasize+num, :] = labelAdd
    datasize += num
    #type 2
    dataAdd[0,0:5] = processa(4, 5)
    dataAdd[0,5:9] = processa(1, 4)
    labelAdd = processa(4, 5)
    dataMat[datasize:datasize+num,0:9] = dataAdd
    dataMat[datasize:datasize+num,9:19] = ageAdd
    labelMat[datasize:datasize+num, :] = labelAdd
    datasize += num
    #type 3
    dataAdd[0,0:5] = processa(4, 5)
    dataAdd[0,5:9] = processa(2, 4)
    labelAdd = processa(4, 5)
    dataMat[datasize:datasize+num,0:9] = dataAdd
    dataMat[datasize:datasize+num,9:19] = ageAdd
    labelMat[datasize:datasize+num, :] = labelAdd
    datasize += num
    #type 4
    dataAdd[0,0:5] = processa(4, 5)
    dataAdd[0,5:9] = processa(3, 4)
    labelAdd = processa(4, 5)
    dataMat[datasize:datasize+num,0:9] = dataAdd
    dataMat[datasize:datasize+num,9:19] = ageAdd
    labelMat[datasize:datasize+num, :] = labelAdd
    datasize += num
    #type 5
    dataAdd[0,0:5] = processa(3, 5)
    dataAdd[0,5:9] = processa(3, 4)
    labelAdd = processa(4, 5)
    dataMat[datasize:datasize+num,0:9] = dataAdd
    dataMat[datasize:datasize+num,9:19] = ageAdd
    labelMat[datasize:datasize+num, :] = labelAdd
    datasize += num
    
    return dataMat, labelMat

def loadDataSet():
    # read data from transition.xlsx
    # store state, action in dataMat, store state' in labelMat
    filename = 'data\\transition.xlsx'
    dirpath = os.path.dirname(os.path.abspath(__file__))
    filename = os.path.join(dirpath, filename)
    wb = load_workbook(filename)
    sheet = wb.active
    maxrow = max_row(sheet)
    index=0
    for row in range(1, max_row(sheet)+1):
        state_num = sheet.cell(row, 1).value #5*1
        action_num = sheet.cell(row, 2).value #4*!
        _state_num = sheet.cell(row, 3).value #5*1
        age_num = math.floor((sheet.cell(row, 4).value-1)/10.0) #10*1
        if not_number(state_num) or not_number(action_num) or not_number(_state_num):
            continue
        elif age_num>=10:
            continue
        else:
            index += 1
    # dataMat:shape[n,2]; labelMat:shape[n,1]
    dataMat = np.zeros(shape=[index, 19], dtype=np.int32)
    labelMat = np.zeros(shape=[index, 5], dtype=np.int32)
    index = 0
    for row in range(1, max_row(sheet)+1):
        state_num = sheet.cell(row, 1).value #5*1
        action_num = sheet.cell(row, 2).value #4*!
        _state_num = sheet.cell(row, 3).value #5*1
        age_num = math.floor((sheet.cell(row, 4).value-1)/10.0) #10*1
        if not_number(state_num) or not_number(action_num) or not_number(_state_num):
            continue
        #do not consider bidges whose age>100
        elif age_num>=10:
            continue
        else:
            dataMat[index, 0:5] = processa(state_num, 5)
            dataMat[index, 5:9] = processa(action_num, 4)
            dataMat[index, 9:19] = processa(age_num, 10)
            labelMat[index, :] = processa(_state_num, 5)
            index += 1
    dataFinal, labelFinal = addData(dataMat, labelMat, 100)
    return dataFinal, labelFinal

def add_layer(inputs, in_size, out_size, n_layer, activation_function=None):
    layer_name = 'layer%s' % n_layer
    with tf.name_scope(layer_name):
        with tf.name_scope('weights'):
            Weights = tf.Variable(tf.random_normal([in_size, out_size]), name='W')
            tf.summary.histogram(layer_name+'/weights',Weights)
        with tf.name_scope('biases'):
            biases = tf.Variable(tf.zeros([1, out_size])+0.1)
            tf.summary.histogram(layer_name+'/biases',biases)
        with tf.name_scope('Wx_plus_b'):
            Wx_plus_b = tf.matmul(tf.to_float(inputs), Weights) + biases
        if activation_function is None:
            outputs = Wx_plus_b
        else:
            outputs = activation_function(Wx_plus_b)
        tf.summary.histogram(layer_name+'/outputs',outputs)
        return outputs

class Predictor:
    def __init__(self, credit, co_phy, lr=0.005, maxCycles=100001, update_cycle=1000, batch_size=200, update_size=20):
        self.credit = credit
        self.lr = lr
        self.co_phy = co_phy
        self.maxCycles = maxCycles
        self.updateCycles = update_cycle
        self.batch_size = batch_size
        self.update_size = update_size
        self.memory = np.zeros(shape=[50*(self.credit+1)+1, 24])

        self._build_net()
        self.sess = tf.Session()
        self.sess.run(tf.global_variables_initializer())
        self.loss = []
        self.loss_update = []

        self.initilize_transition()

        self.generate_data()

    def _build_net(self):
        #placeholder
        self.xs = tf.placeholder(tf.int32, [None, 19])
        self.ys = tf.placeholder(tf.int32, [None, 5])
        #add l1 layer and output layer
        l1 = add_layer(self.xs, 19, 25, n_layer=1, activation_function=tf.nn.relu)
        self.prediction = add_layer(l1, 25, 5, n_layer=2, activation_function=tf.nn.softmax)
        #loss_det, loss_main = self.phy_loss()
        self.cross_entropy = tf.reduce_mean(
            -tf.reduce_sum(tf.to_float(self.ys)*tf.log(self.prediction), reduction_indices=[1])
        )
        #loss_det: only consider state!=4, aciton=0, others are zeors
        #batch size: correspond to "None"
        self.batchSize = tf.placeholder(tf.int32)
        #start and end: 1*None
        self.start = tf.placeholder(tf.int32, [None,1])
        self.end = tf.placeholder(tf.int32, [None, 1])
        length = 5
        #change start and end to None*5
        start = tf.tile(self.start, [1, length])
        end = tf.tile(self.end, [1, length])

        range_vec = tf.expand_dims(tf.range(length), axis=0)
        range_mat = tf.tile(range_vec, [self.batchSize, 1])

        mask_start = tf.where(range_mat>=start, tf.fill([self.batchSize, length], True), tf.fill([self.batchSize, length], False))
        mask_end = tf.where(range_mat<=end, tf.fill([self.batchSize, length], True), tf.fill([self.batchSize, length], False))
        mask = tf.cast(tf.logical_and(mask_start, mask_end), dtype=tf.int32)
        need_seq = tf.multiply(self.prediction, tf.to_float(mask))
        self.loss_phy = tf.reduce_mean(tf.reduce_sum(need_seq, 1))

        self.train_loss = self.co_phy*self.loss_phy+self.cross_entropy

        self.train_step = tf.train.GradientDescentOptimizer(self.lr).minimize(self.train_loss)

        #saver
        self.saver = tf.train.Saver(max_to_keep=5)

    def compute_accuracy(self,v_xs, v_ys):
        y_pre = self.sess.run(self.prediction, feed_dict={self.xs:v_xs})
        correct_prediction = tf.equal(tf.argmax(y_pre,1), tf.argmax(v_ys,1))
        accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
        result = self.sess.run(accuracy, feed_dict={self.xs:v_xs, self.ys:v_ys})
        return result

    def generate_transition(self):
        # set the transition matrix according to the newest transition model
        if not hasattr(self, 'transition'):
            self.transition = np.zeros(shape=[10, 4, 5, 5], dtype=np.float32)
        for year in range(10):
            for action in range(4):
                for state in range(5):
                    x_input = np.zeros(shape=[1, 19], dtype=np.int32)
                    x_input[0,0:5] = processa(state,5)
                    x_input[0,5:9] = processa(action, 4)
                    x_input[0,9:19] = processa(year, 10)
                    self.transition[year, action, state] = self.sess.run(self.prediction, feed_dict={self.xs: x_input})[0]
                    '''
                    if state == 0 and action == 0:
                        self.transition[year, action, state] = [1,0,0,0,0]
                    else:
                        x_input = np.zeros(shape=[1, 19], dtype=np.int32)
                        x_input[0,0:5] = processa(state,5)
                        x_input[0,5:9] = processa(action, 4)
                        x_input[0,9:19] = processa(year, 10)
                        self.transition[year, action, state] = self.sess.run(self.prediction, feed_dict={self.xs: x_input})[0]
                    '''
        #store transition
        #self.save_transition(episode=-1)

    
    def save_transition(self, episode=-1, year=0):
        if episode == -1:
            # initial transition model
            path = 'result\\transition model'
            dirpath = os.path.dirname(os.path.abspath(__file__))
            path = os.path.join(dirpath, path)
            if not os.path.exists(path):
                os.makedirs(path)
            np.save(path+'\\transition.npy', self.transition)
        else:
            # adjusted transition model
            path = 'result\\simulation\\onlineDQN'
            path = path+'\\episode'+str(episode)
            dirpath = os.path.dirname(os.path.abspath(__file__))
            path = os.path.join(dirpath, path)
            if not os.path.exists(path):
                os.makedirs(path)
            np.save(path+'\\year'+str(year)+'-transition.npy', self.transition)
    
    def initilize_transition(self):
        dataMat, labelMat = loadDataSet()
        self.input = dataMat
        self.output = labelMat
        m,n = np.shape(dataMat)
        trainset_len = int(m*0.8)

        #split train and test
        row_indices = np.random.permutation(dataMat.shape[0])
        dataTrain = dataMat[row_indices[0:trainset_len], :]
        dataTest = dataMat[row_indices[trainset_len: m], :]
        labelTrain = labelMat[row_indices[0:trainset_len], :]
        labelTest = labelMat[row_indices[trainset_len: m], :]

        #training set size
        dataset_size, _ = np.shape(dataTrain)

        #start training
        for loop in range(self.maxCycles):
            start = (loop*self.batch_size) % dataset_size
            end = min(start+self.batch_size, dataset_size)
            batchSize = min(end-start, self.batch_size)
            #design start and end vector (batch_size)
            #action=0: state+1~4; action=0: 0~state-1 (include border)
            start_vec = np.zeros(shape=[batchSize,1], dtype=np.int32)
            end_vec = np.zeros(shape=[batchSize,1], dtype=np.int32)
            for i in range(start, end):
                state = deprocessa(dataTrain[i, 0:5], size=5)
                action = deprocessa(dataTrain[i, 5:9], size=4)
                if action==0:
                    start_vec[i-start] = state+1
                    end_vec[i-start] = 4
                else:
                    start_vec[i-start] = 0
                    end_vec[i-start] = state-1
            
            self.sess.run(self.train_step, feed_dict={
            self.xs:dataTrain[start:end], self.ys:labelTrain[start:end], self.start:start_vec, self.end:end_vec, self.batchSize:batchSize})

            if loop%100 == 0:
                total_cross_entropy = self.sess.run(self.cross_entropy, feed_dict={
                    self.xs: dataTrain, self.ys: labelTrain})
                #print(total_cross_entropy)
                self.loss.append(total_cross_entropy)
            
            if loop%10000 == 0 and loop>0:
                path = 'result\\transition model\\loop'+str(loop)
                dirpath = os.path.dirname(os.path.abspath(__file__))
                path = os.path.join(dirpath, path)
                if not os.path.exists(path):
                    os.makedirs(path)
                checkpoint_path = os.path.join(path+'\\training-'+str(loop)+'.ckpt')
                self.saver.save(self.sess, checkpoint_path)
    
        #print accuracy
        self.accuracy = self.compute_accuracy(dataTest, labelTest)
        #print('accuracy'+str(self.compute_accuracy(dataTest, labelTest)))

        #generate transition
        self.generate_transition()
        #self.plot_loss(self.loss)

        #save the transition matrix
        self.save_transition()

    def plot_loss(self,loss):
        plt.plot(np.arange(len(loss)), loss)
        plt.ylabel('Loss')
        plt.xlabel('training steps')
        plt.show()

    def store_transition(self, state, action, state_, age):
        data = np.zeros(shape=[1, 24], dtype=np.int32)
        data[0,0:5] = state[0:5]
        data[0,5:9] = processa(action, 4)
        period = math.floor((age-1)/10.0)
        data[0,9:19] = processa(period, 10)
        data[0,19:24] = state_[0:5]
        self.memory[self.memory_counter, :] = data
        self.memory_counter += 1

    def generate_data(self):
        self.memory_counter = 0
        #generate dummy data based on previous transition model
        #credit=n(model)/n(real)
        #n(real)=50, n(model)=50*credit
        #randomly select 50*credit data into memory
        row_indices = np.random.permutation(self.input.shape[0])
        data_size, _ = np.shape(self.input)
        if 50*self.credit>data_size:
            #select data_size number
            input_store = self.input[row_indices[0:data_size],:]
            output_store = self.output[row_indices[0:data_size],:]
            self.memory[self.memory_counter:self.memory_counter+data_size,:] = np.hstack((input_store, output_store))
            self.memory_counter += data_size
        else:
            #select 50*credit number
            input_store = self.input[row_indices[0:50*self.credit],:]
            output_store = self.output[row_indices[0:50*self.credit],:]
            self.memory[self.memory_counter:self.memory_counter+50*self.credit,:] = np.hstack((input_store, output_store))
            self.memory_counter += 50*self.credit
    
    def update(self):
        input_data = np.hsplit(self.memory[0:self.memory_counter], [19])[0]
        output_data = np.hsplit(self.memory[0:self.memory_counter], [19])[1]
        row_indices = np.random.permutation(input_data.shape[0])
        dataMat = input_data[row_indices, :]
        labelMat = output_data[row_indices, :]

        #training set size
        dataset_size, _ = np.shape(dataMat)

        #start training
        for i in range(self.updateCycles):
            start = (i*self.update_size) % dataset_size
            end = min(start+self.update_size, dataset_size)
            batchSize = min(end-start, self.update_size)
            #design start and end vector (batch_size)
            #action=0: state+1~4; action=0: 0~state-1 (include border)
            start_vec = np.zeros(shape=[batchSize,1], dtype=np.int32)
            end_vec = np.zeros(shape=[batchSize,1], dtype=np.int32)
            for i in range(start, end):
                state = deprocessa(dataMat[i, 0:5], size=5)
                action = deprocessa(dataMat[i, 5:9], size=4)
                if action==0:
                    start_vec[i-start] = state+1
                    end_vec[i-start] = 4
                else:
                    start_vec[i-start] = 0
                    end_vec[i-start] = state-1
            self.sess.run(self.train_step, feed_dict={
            self.xs:dataMat[start:end], self.ys:labelMat[start:end], self.start:start_vec, self.end:end_vec, self.batchSize:batchSize})
            '''
            self.sess.run(self.train_step, feed_dict={
            self.xs:dataMat[start:end], self.ys:labelMat[start:end]})
            '''

            if i%100 == 0:
                total_cross_entropy = self.sess.run(self.cross_entropy, feed_dict={
                    self.xs: dataMat, self.ys: labelMat})
                #print(total_cross_entropy)
                self.loss_update.append(total_cross_entropy)
        
        self.generate_transition()
        #self.plot_loss(self.loss_update)
    
    def _import_variable(self):
        path = 'result\\transition model\\loop100000'
        dirpath = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(dirpath, path)
        moudke_file = tf.train.latest_checkpoint(path)
        self.saver.restore(self.sess, moudke_file)

    def reset(self):
        # for every training episode
        # reset the memory, network parameter, transition matrix
        # reset the network parameter
        self._import_variable()
        # reset the transition matrix
        path = 'result\\transition model\\transition.npy'
        dirpath = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(dirpath, path)
        self.transition = np.load(path)
        # reset the memory
        self.generate_data()