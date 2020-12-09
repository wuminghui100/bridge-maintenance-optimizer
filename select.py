from openpyxl import load_workbook
from os import path
import math

def not_number(s):
    try:
        float(s)
        return False
    except ValueError:
        pass
 
    try:
        import unicodedata
        unicodedata.numeric(s)
        return False
    except (TypeError, ValueError):
        pass
 
    return True

def max_row(sheet):
    for i in range(1, sheet.max_row+2):
        if sheet.cell(i, 1).value == None:
            return i-1

def state_trans(state_raw):
    if state_raw==9:
        state=4
    elif state_raw>=8:
        state=3
    elif state_raw>=5:
        state=2
    elif state_raw>=2:
        state=1
    else:
        state=0
    return state

# a function to select eligible bridges
# if not eligible, return False; otherwise return True
def check_meet_requirment(sheet, row):
    '''
    #exclude reconstructed bridges
    if sheet.cell(row, 106).value!=0:
        return False
    #only include local bridges
    if sheet.cell(row, 26).value!=9 and sheet.cell(row, 26).value!=19:
        return False
    '''
    #only include low ADT bridges
    if sheet.cell(row, 30).value== None or not_number(sheet.cell(row, 30).value) or sheet.cell(row, 30).value>5000:
        return False
    '''
    #filter construction time:
    if sheet.cell(row, 27).value== None or not_number(sheet.cell(row, 27).value) or int(sheet.cell(row, 27).value)<1960:
        return False
    '''
    #filter structure type-only include concrete
    if sheet.cell(row, 48).value!=1:
        return False
    #file deck type-only include concrete
    if sheet.cell(row, 107).value!=1:
        return False
    #selcet inspection interval-only include 24months
    if sheet.cell(row, 86).value!=24:
        return False
    #select surface type-only include concrete
    if sheet.cell(row, 108).value == None or not_number(sheet.cell(row, 108).value) or int(sheet.cell(row, 108).value)>4:
        return False
    #exclude false CR(N and 0):
    if not_number(sheet.cell(row, 67).value) or int(sheet.cell(row, 67).value)==0:
        return False
    return True

#state from 1~8, _state from 2~9
def check_action(state, _state):
    BIM = [[1,0,0,0,0,0,0,0],
           [1,1,0,0,0,0,0,0],
           [3,2,1,0,0,0,0,0],
           [3,3,2,1,0,0,0,0],
           [3,3,2,2,1,0,0,0],
           [3,3,2,2,2,1,0,0],
           [3,3,2,2,2,1,1,0],
           [3,3,2,2,2,1,1,1]]
    raw = 8-state
    col = 9-_state
    return BIM[raw][col]

#a funtion to write state, action, _state in another file
#sheet 1 provides state, action in line row;
#sheet 2 provides _state;
#write on sheet 3 in line _row
def write_file(sheet1, sheet2, sheet3, row, _row):
    state_raw = sheet1.cell(row, 67).value
    if state_raw == None or not_number(state_raw):
        return False
    state = state_trans(state_raw)
    
    current_year = 2000+num
    # only when current year shares same parity with last inspection year
    # the record is valid
    last_inspection = sheet1.cell(row, 85).value
    if last_inspection==None or not_number(last_inspection) or current_year%2 != last_inspection%2:
        return False
    id = sheet1.cell(row, 2).value
    for i in range(1, max_row(sheet2)+1):
        if sheet2.cell(i, 2).value==id:
            
            _state_raw = sheet2.cell(i, 67).value
            if _state_raw == None or not_number(_state_raw):
                continue
            #improvement cost
            improve = sheet1.cell(row,93).value
            if improve == None or not_number(improve):
                continue
            #determine action
            if state_raw<_state_raw:
                action = check_action(state_raw, _state_raw)
            elif state_raw == _state_raw and improve >0:
                action = 1
            else:
                action = 0
            
            #calcute bridge age
            if sheet1.cell(row, 27).value == None or not_number(sheet1.cell(row, 27).value):
                return False
            else:
                construct_year = int(sheet1.cell(row, 27).value)
            age = current_year - construct_year
            age = 10*int(math.ceil(age/10.0))
            if age>100:
                continue
            #write in sheet3
            _state = state_trans(_state_raw)
            sheet3.cell(_row, 1).value=state
            sheet3.cell(_row, 2).value=action
            sheet3.cell(_row, 3).value=_state
            sheet3.cell(_row, 4).value= age
            return True
    return False

#iteration of each row in sheet1
def run(sheet1, sheet2, sheet3):
    _row = max_row(sheet3)+1
    for num in range(1, sheet1.max_row):
        if check_meet_requirment(sheet1, num):
            print(_row)
            result = write_file(sheet1, sheet2, sheet3, num, _row)
            if result:
                _row += 1


#absolute path to active files
global num
for num in range(-8,18,1):
    #file1: state and action
    if num<0:
        _filename1 = 'PA'+str(100+num)+'.xlsx'
    elif(num<10):
        _filename1 = 'PA0'+str(num)+'.xlsx'
    else:
        _filename1 = 'PA'+str(num)+'.xlsx'
    dir_path = path.dirname(path.abspath(__file__))
    filename1 = path.join(dir_path,_filename1)
    wb1 = load_workbook(filename1)
    sheet1 = wb1.active
    #file2: _state
    _num = num+2
    if _num<0:
        _filename2 = 'PA'+str(100+_num)+'.xlsx'
    elif(_num<10):
        _filename2 = 'PA0'+str(_num)+'.xlsx'
    else:
        _filename2 = 'PA'+str(_num)+'.xlsx'
    dir_path = path.dirname(path.abspath(__file__))
    filename2 = path.join(dir_path,_filename2)
    wb2 = load_workbook(filename2)
    sheet2 = wb2.active
    #file3: write in this file
    _filename3 = 'transition.xlsx'
    dir_path = path.dirname(path.abspath(__file__))
    filename3 = path.join(dir_path,_filename3)
    wb3 = load_workbook(filename3)
    sheet3 = wb3.active
    #run the main function
    run(sheet1, sheet2, sheet3)
    wb3.save(filename3)

print("done")